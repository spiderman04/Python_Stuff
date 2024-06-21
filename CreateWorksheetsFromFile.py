import openpyxl
import os
import sys
from openpyxl import Workbook as Workbook
from openpyxl.styles import Alignment, Font, PatternFill

#define script constants and data structures
EXCEL_CONFIG_FILENAME = "CreateWorksheets.xlsx"  # flexibility to rename output file
TABLENAME_FILE = '\Inputs\Listing.txt'
OUTPUT_DIRNAME = '\\Inputs\\Tables\\'
READ_CONFIG_FILENAME = '\Inputs\querybuilder.properties'
EXCEL_MAX_ROWS = 1048576
EXCEL_MAX_COLS = 16384
SEPERATOR_CHAR = '='    # used in .properties file
CONFIG_KEYS = {}
NEW_LINE_CHAR = '\n'

def parseDataFile(file_to_read):

    """
    Descriptyion
    ------------
    Used to count the configuration keys which match the list above
    and have TRUE values. This number is used for the highlighting logic
    to group the coloring scheme across different table names.

    :param file_to_read: the input file to parse/scan

    :return: None
    """
    try:
        with open(file_to_read) as f:
            for line in f:
                # skip the lines that begin with '#"
                if not line.startswith('#', 0, len(line)):
                    if SEPERATOR_CHAR in line:
                        # find the key/value pair by splitting the string
                        name, value = line.split(SEPERATOR_CHAR, 1)
                        CONFIG_KEYS[name.strip()] = value.strip()

        # loop thru the 'Keys' dictionary and clean the lines with '#' character
        for k,v in CONFIG_KEYS.items():
            # find the position of the '#' character
            pos = v.find('#')
            if pos > -1:
                # strip away the comment line and replace the Key dictionary with new value
                CONFIG_KEYS[k] = str(v[0:pos-1])
            else:
                continue

    except Exception as e:
         print("An exception has occured: " + str(e))

def setup_excel_headers(excel_file):
    #open the new file for adding Rows to
    work_book = openpyxl.load_workbook(excel_file)
    ws = work_book['Config']

    #clear out any previous data in the file
    ws.delete_rows(1, EXCEL_MAX_ROWS)
    ws.delete_cols(1, EXCEL_MAX_COLS)

    #adjust all headers Font, Size, Bold and Centered
    for i in range(1,8):
        ws.cell(1,i).font = Font(name='Calibri', size=14, bold=True)
        ws.cell(1,i).alignment = Alignment(horizontal='center', vertical='center', wrapText=False)

    #save final edits
    work_book.save(excel_file)
    work_book.close()

def main():
    file_path = os.getcwd()
    print("Reading input from...." + file_path)

    table_name_file_path = file_path + TABLENAME_FILE
    excel_config_file_path = file_path + OUTPUT_DIRNAME + EXCEL_CONFIG_FILENAME

    #open the file to read in the Master Table names file
    try:
        openFd = open(table_name_file_path, 'r')
        table_names = openFd.readlines()
        openFd.close()
    except OSError as e:
        print(sys.stderr, str(e))

    #create a primary EMPTY config Excel file
    wb = openpyxl.workbook.Workbook()
    ws = wb.active
    ws.title = "Config"
    wb.save(excel_config_file_path)

    for table_name in table_names:
        if table_name != '\n':
            print("query for table.... " + table_name)
            new_sheet = wb.create_sheet(table_name)
            new_sheet.title = table_name
            
    #save final edits
    wb.save(excel_config_file_path)
