
"""
This program will create the main Excel configuration for automating all SQL
wqueries

References:
    Openpyxl Python Library
    https://openpyxl.readthedocs.io/en/latest/changes.html#id30
"""
import sys

import openpyxl
import os
import pyodbc
import sys
from openpyxl import Workbook as Workbook
from openpyxl.styles import Alignment, Font, PatternFill

#define script constants and data structures
EXCEL_CONFIG_FILENAME = "ConfigBuilder.xlsx"  # flexibility to rename output file
TABLENAME_FILE = '\Inputs\MasterTable.txt'
OUTPUT_DIRNAME = '\\Inputs\\Tables\\'
READ_CONFIG_FILENAME = '\Inputs\querybuilder.properties'
EXCEL_MAX_ROWS = 1048576
EXCEL_MAX_COLS = 16384
SEPERATOR_CHAR = '='    # used in .properties file
CONFIG_KEYS = {}
NEW_LINE_CHAR = '\n'

#Define a control list to be compared with properties file
VALID_GRAMMER_LIST = [
    "BLD_QRY_DESCRIBE","BLD_QRY_COUNT","BLD_QRY_DUPLICATE","BLD_QRY_DAYWISE","BLD_QRY_NULL_COLUMNS",
    "BLD_QRY_SAMPLE_DATA","BLD_QRY_EDP_NULL_CHECK","BLD_QRY_MINUS_AB","BLD_QRY_MINUS_BA",
    "BLD_QRY_HIVE_INCREMENTAL","BLD_QRY_HIVE_HISTORICAL","BLD_QRY_IS_NOT_NULL"
]

def BuildDuplicateQuery(SQLQuery, table_name):
    snowflake_dsn = CONFIG_KEYS['DSN_SNOWFLAKE']
    conn = pyodbc.connect("DSN=" + snowflake_dsn)
    Cursor = conn.cursor()

def countConfigKeys():
    """
    Used to count the configuration keys which match the grammer list above AND have a TRUE value. This
    number is used for the highlighting logic to group the colorin gsheme across different table names.

    :return: key_count: the count of valid keys = TRUE
    """
    key_count = 0
    for k,v in CONFIG_KEYS.items():
        pos = v.lower().find("true")

        # only count the keys in the vlaid grammer list and equakl TRUE
        if (pos > -1) and (k in VALID_GRAMMER_LIST):
            key_count += 1
        else:
            continue

    return key_count

def parsePropertiesFile(file_to_read):
    """
    Descriptyion
    ------------
    Used to count the configuration keys which match the list above
    and have TRUE values. This number is used for the highlighting logic
    to group the coloring scheme across different table names.

    :param file_to_read: the input file to parse/scan

    :return: None
    """
    try:
        with open(file_to_read) as f:
            for line in f:
                # skip the lines that begin with '#"
                if not line.startswith('#', 0, len(line)):
                    if SEPERATOR_CHAR in line:
                        # find the key/value pair by splitting the string
                        name, value = line.split(SEPERATOR_CHAR, 1)
                        CONFIG_KEYS[name.strip()] = value.strip()

        # loop thru the 'Keys' dictionary and clean the lines with '#' character
        for k,v in CONFIG_KEYS.items():
            # find the position of the '#' character
            pos = v.find('#')
            if pos > -1:
                # strip away the comment line and replace the Key dictionary with new value
                CONFIG_KEYS[k] = str(v[0:pos-1])
            else:
                continue

    except Exception as e:
        print("An exception has occured: " + str(e))

def setup_excel_headers(excel_file):
    #open the new file for adding Rows to
    work_book = openpyxl.load_workbook(excel_file)
    ws = work_book['Config']

    #clear out any previous data in the file
    ws.delete_rows(1, EXCEL_MAX_ROWS)
    ws.delete_cols(1, EXCEL_MAX_COLS)

    #write the header once
    cell_obj = ws.cell(row=1, column=1)
    cell_obj.value = "S.NO."
    cell_obj = ws.cell(row=1, column=2)
    cell_obj.value = "SQL Query"
    cell_obj = ws.cell(row=1, column=3)
    cell_obj.value = "SHEETNAME"
    cell_obj = ws.cell(row=1, column=4)
    cell_obj.value = "HEADING"
    cell_obj = ws.cell(row=1, column=5)
    cell_obj.value = "TABLE NAME"
    cell_obj = ws.cell(row=1, column=6)
    cell_obj.value = "DSN CONNECTOR"
    cell_obj = ws.cell(row=1, column=7)
    cell_obj.value = "SKIP ROW?"

    #change background color to orange
    ws["A1"].fill = PatternFill("solid", start_color="FFFFC000")
    ws["B1"].fill = PatternFill("solid", start_color="FFFFC000")
    ws["c1"].fill = PatternFill("solid", start_color="FFFFC000")
    ws["d1"].fill = PatternFill("solid", start_color="FFFFC000")
    ws["e1"].fill = PatternFill("solid", start_color="FFFFC000")
    ws["f1"].fill = PatternFill("solid", start_color="FFFFC000")
    ws["g1"].fill = PatternFill("solid", start_color="FFFFC000")

    #adjust all headers Font, Size, Bold and Centered
    for i in range(1,8):
        ws.cell(1,i).font = Font(name='Calibri', size=14, bold=True)
        ws.cell(1,i).alignment = Alignment(horizontal='center', vertical='center', wrapText=False)

    #fix last row format??
    ws.cell(ws.max_row, 1).alignment = Alignment(horizontal='center', vertical='center', wrapText=False)

    # how many TRUE keys are read from file
    query_count = countConfigKeys()

    # highlight alternate rows
    group_color = 1
    for i in range(2, ws.max_row+1):
        for j in range(1, ws.max_column+1):
            if group_color % 2 ==0:
                cobj = ws.cell(row=i, column=j)
                cobj.fill = PatternFill("solid", start_color='C5D9F1') # light blue
            else:
                cobj = ws.cell(row=i, column=j)
                cobj.fill = PatternFill("solid", start_color='D8E4BC')  # light green
        # increment each block for each table
        if i % (query_count+1) == 0:
            group_color += 1
    #adjust column widths
    #reference: https://python-bloggers.com/2023/05/how-to-automatucally-adjut-excel-column-width-in-openpyxl
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        if column_letter != 'B':
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except: # trap length errors
                    pass

            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        else:
            # set column B to a decent size, say 100 pixels
            ws.column_dimensions['B'].width = 100.00

    #save final edits
    work_book.save(excel_file)
    work_book.close()

def write_excel_row(ws_obj, table_name, tc_counter, qry_type):
    try:
        row_num = ws_obj.max_row
        # check the query tpye for TRUE for writing
        if CONFIG_KEYS[qry_type].lower() == 'true':
            #ws_obj.cell(row_num, 1).align
            row_num = ws_obj.max_row + 1

            #q,h = get_query_type(qry_type, table_name)

            cell_obj = ws_obj.cell(row=row_num, column=1)
            cell_obj.value = tc_counter
            cell_obj = ws_obj.cell(row=row_num, column=2)
            cell_obj.value = "select count(*) from carrwdev.X" + str(tc_counter)
            cell_obj = ws_obj.cell(row=row_num, column=3)
            cell_obj.value = table_name
            cell_obj = ws_obj.cell(row=row_num, column=4)
            cell_obj.value = str("Header Name ") + str(tc_counter)
            cell_obj = ws_obj.cell(row=row_num, column=5)
            cell_obj.value = table_name
            cell_obj = ws_obj.cell(row=row_num, column=6)
            if (qry_type == 'BLD_QRY_HIVE_INCREMENTAL' or qry_type == 'BLD_QRY_HIVE_HISTORICAL'):
                cell_obj.value = CONFIG_KEYS['DSN_HIVE']
            elif (qry_type == 'DSN_SNOWFLAKE'):
                cell_obj.value = CONFIG_KEYS['DSN_SNOWFLAKE']
            else:
                cell_obj.value = "Unknown"
            # skip row column string
            cell_obj = ws_obj.cell(row=row_num, column=7)
            cell_obj.value = "N"

            row_num += 1
            tc_counter += 1

    except KeyError:
        print("ERROR: Do not recognize query type: " + qry_type)
        pass

    return row_num, tc_counter
def populate_excel(workbook_obj, worksheet_obj, excel_config_file_path, table_name):

    # write each Row based on config file settings
    r,c = write_excel_row(worksheet_obj, table_name, 1, "BLD_QRY_DESCRIBE")
    r,c = write_excel_row(worksheet_obj, table_name, c, "BLD_QRY_COUNT")
    #workbook_obj.save(excel_config_file_path)

def main():
    file_path = os.getcwd()
    print("Reading input from...." + file_path)

    read_config_file_path = file_path + READ_CONFIG_FILENAME
    parsePropertiesFile(read_config_file_path)

    table_name_file_path = file_path + TABLENAME_FILE
    excel_config_file_path = file_path + OUTPUT_DIRNAME + EXCEL_CONFIG_FILENAME

    #open the file to read in the Master Table names file
    try:
        openFd = open(table_name_file_path, 'r')
        table_names = openFd.readlines()
        openFd.close()
    except OSError as e:
        print(sys.stderr, str(e))

    #create a primary EMPTY config Excel file
    wb = openpyxl.workbook.Workbook()
    ws = wb.active
    ws.title = "Config"
    wb.save(excel_config_file_path)
    setup_excel_headers(excel_config_file_path)

    for table_name in table_names:
        if table_name != '\n':
            print("query for table.... " + table_name)
            populate_excel(wb, ws, excel_config_file_path, table_name.replace('\n', ''))

    #save final edits
    wb.save(excel_config_file_path)


if __name__ == '__main__':
    main()
